from openpyxl import Workbook
from PIL import Image
from numpy import asarray
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def make_sheet_from_pic(pic_path, wb_path, size=[60, 30]):
    """Converts an image file to an Excel Worksheet; Saves Workbook file

    Parameters:
    pic_path (str): the path of your image file
    wb_path(str):   the path of the new Workbook; should end in .xlsx
    size(list):     what size you want the image to be converted to

    Returns:
    None

   """
    # Make Workbook and Worksheet
    wb = Workbook()
    ws = wb.active
    # Load image
    img = Image.open(pic_path)
    # Shrink down to size
    img.thumbnail(size, Image.ANTIALIAS)
    # Convert image to numpy array
    data = asarray(img)
    # Fill in sheet cells with matching rgb
    for row in range(data.shape[0]):
        for col in range(data.shape[1]):
            rgb = data[row][col]
            colour = '{0:02X}{1:02X}{2:02X}'.format(*rgb)
            fill = PatternFill(patternType="solid", fgColor=colour, bgColor=colour)
            ws.cell(row+1, col+1).fill = fill
    # Change sheet column widths to make cells square
    for i in range(1, data.shape[1]+1):
        ws.column_dimensions[get_column_letter(i)].width = 2.7
    # Save Workbook
    wb.save(wb_path)